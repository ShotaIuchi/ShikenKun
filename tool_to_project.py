#!python3
# coding: utf-8


import shutil
import openpyxl
import json


with open('settings.json', 'r') as fp:
    jData = json.load(fp)
    src_row_start           = jData["src"]["row_start"]
    src_premise_start       = jData["src"]["premise_start"]
    src_premise_end         = jData["src"]["premise_end"]
    src_procedure_start     = jData["src"]["procedure_start"]
    src_procedure_end       = jData["src"]["procedure_end"]
    src_confirmation_start  = jData["src"]["confirmation_start"]
    src_confirmation_end    = jData["src"]["confirmation_end"]
    dst_row_start           = jData["dst"]["row_start"]
    dst_premise             = jData["dst"]["premise"]
    dst_procedure           = jData["dst"]["procedure"]
    dst_confirmation        = jData["dst"]["confirmation"]


def main():
    src = openpyxl.load_workbook('./source_test_data.xlsx')
    dst = openpyxl.load_workbook('./template.xlsx')

    srcSheet = src[src.sheetnames[0]]
    dstSheet = dst[dst.sheetnames[0]]

    dstColumn = dst_row_start

    for row in range(src_row_start, srcSheet.max_row):
        """ 
        if (srcSheet.cell(row=row, column=1).value >= 1):
            continue
        """
        value = ''
        for column in range(src_premise_start, src_premise_end):
            cell = srcSheet.cell(row=row, column=column)
            if (cell.value != None):
                value += '・'
                value += cell.value
                value += '\r\n'
        dstSheet.cell(row=dstColumn, column=dst_premise).value = value
        dstSheet.cell(row=dstColumn, column=dst_premise).alignment = openpyxl.styles.Alignment(wrapText=True)

        value = ''
        num = 1
        for column in range(src_procedure_start, src_procedure_end):
            cell = srcSheet.cell(row=row, column=column)
            if (cell.value != None):
                value += str(num)
                value += '.'
                value += cell.value
                value += '\r\n'
                num += 1
        dstSheet.cell(row=dstColumn, column=dst_procedure).value = value
        dstSheet.cell(row=dstColumn, column=dst_procedure).alignment = openpyxl.styles.Alignment(wrapText=True)
        
        value = ''
        for column in range(src_confirmation_start, src_confirmation_end):
            cell = srcSheet.cell(row=row, column=column)
            if (cell.value != None):
                value += '・'
                value += cell.value
                value += '\r\n'
        dstSheet.cell(row=dstColumn, column=dst_confirmation).value = value
        dstSheet.cell(row=dstColumn, column=dst_confirmation).alignment = openpyxl.styles.Alignment(wrapText=True)
        
        dstColumn += 1
    dst.save('dust_test_data.xlsx')


if __name__ == '__main__':
    print('[S]tool_to_project')
    main()
    print('[E]tool_to_project')
