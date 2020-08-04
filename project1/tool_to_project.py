#!python3
# coding: utf-8


import shutil
import openpyxl


def main():
    src = openpyxl.load_workbook('../source_test_data.xlsx')
    dst = openpyxl.load_workbook('./template.xlsx')

    srcSheet = src[src.sheetnames[0]]
    dstSheet = dst[dst.sheetnames[0]]

    dstColumn = 3

    for row in range(2, srcSheet.max_row):

        if (srcSheet.cell(row=row, column=1).value >= 1):
            continue

        value = ''
        for column in range(2, 5):
            cell = srcSheet.cell(row=row, column=column)
            if (cell.value != None):
                value += '・'
                value += cell.value
                value += '\r\n'
        dstSheet.cell(row=dstColumn, column=3).value = value
        dstSheet.cell(row=dstColumn, column=3).alignment = openpyxl.styles.Alignment(wrapText=True)

        value = ''
        num = 1
        for column in range(6, 9):
            cell = srcSheet.cell(row=row, column=column)
            if (cell.value != None):
                value += str(num)
                value += '.'
                value += cell.value
                value += '\r\n'
                num += 1
        dstSheet.cell(row=dstColumn, column=4).value = value
        dstSheet.cell(row=dstColumn, column=4).alignment = openpyxl.styles.Alignment(wrapText=True)
        
        value = ''
        for column in range(10, 13):
            cell = srcSheet.cell(row=row, column=column)
            if (cell.value != None):
                value += '・'
                value += cell.value
                value += '\r\n'
        dstSheet.cell(row=dstColumn, column=5).value = value
        dstSheet.cell(row=dstColumn, column=5).alignment = openpyxl.styles.Alignment(wrapText=True)
        
        dstColumn += 1
    dst.save('dust_test_data.xlsx')


if __name__ == '__main__':
    print('[S]tool_to_project')
    main()
    print('[E]tool_to_project')
